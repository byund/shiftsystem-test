#auxiliary functions for this program
from math import floor, ceil
import re
import datetime
from openpyxl import Workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

def getListOfWorkers(workbook):
	staffSheet = workbook['CarlTech Staff List']
	#Get list of all the workers from the ITS Schedule
	workerList = []
	#staffNumber = staffSheet.max_row #subtract 1 to get actual number of workers, because the first row will always be the column headers for titles and whatnot.
	staffNumber = input("What is the last row on which there is a Worker in the sheet?" )
	print ("ShiftSystem has detected that your sheet of workers contains " + str(staffNumber) + " rows. If this is too high, the import will fail.")
	for row in range(2, int(staffNumber)+1): #for every worker (range must go to staffNumber+1 because )
		firstName = staffSheet.cell(row = row, column = 1).value
		lastName = staffSheet.cell(row = row, column = 2).value
		username = staffSheet.cell(row = row, column = 5).value
		colleagueID = staffSheet.cell(row = row, column = 6).value
		workerList.append((firstName, (lastName, username, colleagueID)))
	listofWorkers = dict(workerList) #workerList is a tuple; we're making a dictionary out of it.

	return listofWorkers


def dateParser(dateString): #input should be a string in format mm/dd/yy or mm-dd-yy
	values = re.split('\D', dateString, 2) #split string at most twice (into 3 resulting values), based on the decimals. 
	date = datetime.date(int(values[2]), int(values[0]), int(values[1]))
	return date

def getStudent(workerList, workerName): #given the student name in a cell, finds the full name of the worker for Shiftplanning.
#when called, workerList will be a dictionary of workers (See ScheduleImporter.py, under the section commented "Get List of All Workers")
	
	if workerName == "": #Skips blank cells.
		return ""
	else:
		info = workerList.get(workerName, ("No", "Worker", 0000)) #info is a tuple of form (lastname, username, colleagueID)
		firstName = workerName
		lastName = info[0]
		username = info[1]
		colleagueID = info[2]

	#	if firstName == "Sophia M." or firstName == "Sophia B.": #Built in Sophia Processing!
	#		firstName = "Sophia"
	#For cases where there are duplicate first names, they will be distinguished by having the first letter of their last name, e.g. Sophia M. and Sophia B.
	#We will split these strings by space and then give them the first name.
		if workerName:
			firstName = workerName.split()[0]
			colleagueID = int(colleagueID)
		return (firstName, lastName, username, colleagueID)

def getDay(row, col): #returns Day as string from the excel spreadsheet. This might actually be a redundant method but hey, it works!

	
	Daynumber = ceil(col/2)
	
	#now check if you're at midnight on Sunday
	if (col == 14 and row >= 51):
		Daynumber = 1 #roll over to Monday

	elif (col % 2 == 0 and col >= 48):
		Daynumber += 1 
	
	if Daynumber == 1:
		Day = "Monday"
	elif Daynumber == 2:
		Day = "Tuesday"
	elif Daynumber == 3:
		Day = "Wednesday"
	elif Daynumber == 4:
		Day = "Thursday"
	elif Daynumber == 5:
		Day = "Friday"
	elif Daynumber == 6:
		Day = "Saturday"
	elif Daynumber == 7:
		Day = "Sunday"




	return Day

def getLocation(col): #For a given day, if the carltech is in the left column they're at the Helpdesk (CarlTech) and for the right column they're the Library (Libe CarlTech). Modify this code if position names change in the future.
	if (col%2) != 0:
		location = "CMC"
	else:
		location = "ResearchIT"

	return location

def testGetShiftTime(shiftTimeString):
	shiftRegex = re.compile('\d?\d:\d*') #This regex matches anything that looks like it's a shift time. It searches for anything formatted as x:00
	if shiftTimeString.match(shiftRegex) == true:
		print (shiftTimeString.match.group(1))


def getShiftTime(row, col): #gets shift time given a certain cell.
	


	day = ceil(col/2)
	time = ""
	if row > 30 and day < 5: #if we're at 6:00pm on Mon-Thur
		if row > 30 and row < 33:
			time = "18:00"

		elif row > 33 and row < 36:
			time = "19:00"

		elif row > 36 and row < 39:
			time = "20:00"

		elif row > 39 and row < 42:
			time = "21:00"

		elif row > 42 and row < 45:
			time = "22:00"
		elif row > 45 and row < 48:
			time = "23:00"
		elif row > 48 and row < 51:
			time = "00:00"

	elif row > 42 and day > 4 and day < 7: #if we're past 10:00pm on a friday or saturday
		return

	elif day == 1 or day == 3: #mon/wed morning schedule
		if row > 2 and row < 6:
			time = "8:00"
		elif row > 6 and row < 10:
			time = "9:45"
		elif row >10 and row < 14:
			time = "11:05"
		elif row > 14 and row < 18:
			time = "12:25"
		elif row > 18 and row < 22:
			time = "13:45"
		elif row > 22 and row < 26:
			time = "15:05"
		elif row > 26 and row < 30:
			time = "16:25"

	elif day == 2 or day == 4: #tue/thurs morning schedule
		if row > 2 and row < 6:
			time = "8:00"
		elif row > 6 and row < 10:
			time = "10:05"
		elif row >10 and row < 14:
			time = "12:00"
		elif row > 14 and row < 18:
			time = "13:10"
		elif row > 18 and row < 22:
			time = "15:05"
		elif row > 26 and row < 30:
			time = "17:00"

	elif day == 5: #The friday schedule
		if row > 2 and row < 6:
			time = "8:00"
		elif row > 6 and row < 10:
			time = "9:35"
		elif row >10 and row < 14:
			time = "10:45"
		elif row > 14 and row < 18:
			time = "11:55"
		elif row > 18 and row < 22:
			time = "13:05"
		elif row > 22 and row < 26:
			time = "14:15"
		elif row > 26 and row < 30:
			time = "15:25"
		elif row > 30 and row < 33:
			time = "16:30"
		elif row > 33 and row < 36:
			time = "18:00"
		elif row > 36 and row < 39:
			time = "19:00"
		elif row > 39 and row < 42:
			time = "20:00"

	elif day == 6: #the saturday schedule
		if row > 2 and row < 6:
			time = "10:00"
		elif row > 6 and row < 10:
			time = "11:00"
		elif row >10 and row < 14:
			time = "12:00"
		elif row > 14 and row < 18:
			time = "13:00"
		elif row > 18 and row < 22:
			time = "14:00"
		elif row > 22 and row < 26:
			time = "15:00"
		elif row > 26 and row < 30:
			time = "16:00"
		elif row > 30 and row < 33:
			time = "17:00"
		elif row > 33 and row < 36:
			time = "18:00"
		elif row > 36 and row < 39:
			time = "19:00"
		elif row > 39 and row < 42:
			time = "20:00"

	elif day == 7: #the sunday schedule
		if row > 2 and row < 6:
			time = "10:00"
		elif row > 6 and row < 10:
			time = "11:00"
		elif row >10 and row < 14:
			time = "12:00"
		elif row > 14 and row < 18:
			time = "13:00"
		elif row > 18 and row < 22:
			time = "14:00"
		elif row > 22 and row < 26:
			time = "15:00"
		elif row > 26 and row < 30:
			time = "16:00"
		elif row > 30 and row < 33:
			time = "17:00"
		elif row > 33 and row < 36:
			time = "18:00"
		elif row > 36 and row < 39:
			time = "19:00"
		elif row > 39 and row < 42:
			time = "20:00"

		elif row > 42 and row < 45:
			time = "21:00"
		elif row > 45 and row < 48:
			time = "22:00"
		elif row > 48 and row < 51:
			time = "23:00"
		elif row > 51:
			time = "00:00"
	return time



