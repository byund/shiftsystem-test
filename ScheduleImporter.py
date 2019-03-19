#This program will take in the Excel spreadsheet that ITS uses and output a formatted spreadsheet that can then be imported into Humanity as is.
#Required libraries: openpyxl and everything that comes with it.

from openpyxl import Workbook
from openpyxl import load_workbook
import AuxiliaryFunctions
import re
from datetime import datetime, timedelta

#TODO: Be able to toggle between Library shifts or helpdesk shifts only or both.

#load in workbook
filename = input("Place ITS Schedule Spreadsheet in the same folder as this program, then type its filename here (e.g Schedule.xlsx): ")
wb = load_workbook(filename)
scheduleSheet = wb.get_sheet_by_name('CarlTech Schedule')
staffSheet = wb.get_sheet_by_name('CarlTech Staff List')


#Get list of all the workers
workerList = []
staffNumber = staffSheet.max_row #subtract 1 to get actual number of workers, because the first row will always be the column headers for titles and whatnot.
for row in range(2, staffNumber+1): #for every worker (range must go to staffNumber+1 because )
	firstName = staffSheet.cell(row = row, column = 1).value
	lastName = staffSheet.cell(row = row, column = 2).value
	workerList.append((firstName,lastName))
listofWorkers = dict(workerList) #workerList is a tuple; we're making a dictionary out of it.

#create the output workbook
outputWb = Workbook()
outputSheet = outputWb.active
outputWb.save('HumanityScheduleImport.xlsx') #note: this will overwrite existing files without warning.

#populate column headers
outputSheet['A1'] = 'name'
outputSheet['B1'] = 'position'

startDate = AuxiliaryFunctions.dateParser(input("When does the term start? (mm/dd/yyyy): ")) #

for col in range(3,10): #populates 7 Columns (C through I) with the dates of the first week of work. 
	dateString = startDate.strftime('%m/%d/%y') #get start date as a string
	outputSheet.cell(row = 1, column = col, value =  dateString)
	startDate = startDate + timedelta(days=1)


#Now to populate the output spreadsheet with worker names, positions, and shift times.
timeRegex = re.compile('[^a-zA-z.\s]') #This regex matches any string that contains a number i.e a non-name cell.
#shiftTime = "8:00am-9:00am" #initalizing shiftTime just because
for col in range(1,15): #this corresponds to columns A through N
	for row in range(2,53): #this corresponds to the maximum row that the schedule spreadsheet has. Modify as necessary
		workerName = scheduleSheet.cell(row = row, column = col).value
		if workerName != None:
			#workerName = str(workerName) #just to make sure you're dealing with a string....
			if timeRegex.search(workerName) == None: #check that you're not actually dealing with a blank cell or a Time Cell, using the regex initialized above.
				name = AuxiliaryFunctions.getStudent(listofWorkers, workerName)
				position = AuxiliaryFunctions.getPosition(col)
				day = AuxiliaryFunctions.getDay(col)
				outputSheet.append((name, position)) #add the name and position for the worker.
				row_count = outputSheet.max_row
				if shiftTime == "12:00am-1:00am": #For those cases where the time actually stretches into the next day, we increment forward by one accordingly, unless it's at the end of the week (a sunday), in which case we set the day back to Monday.
					day += 1
					if day > 9:
						day = 3
				outputSheet.cell(row = row_count, column = day, value = shiftTime) #add the shift time for the worker in question.
			else: #if you're at a time cell, we will now extract the shiftTime.
				cellValue = outputSheet.cell(row = row, column = row)
				shiftTime = AuxiliaryFunctions.getShiftTime(row, col)

#Having populated the week, we expand it out to the end of term.
endDate = AuxiliaryFunctions.dateParser(input("When does the term end? (mm/dd/yyyy): ")) #
col = 9 #This corresponds to column I, the Sunday of the first week.
lastEnteredDate = datetime.strptime(outputSheet.cell(row = 1, column = col).value, '%m/%d/%y').date() #get last date that the sheet currently contains.
maxRow = outputSheet.max_row

#populates column headers properly
while endDate != lastEnteredDate: #This loop will populate all the way to the last day of term, based on the date entered by the user.
	col+=1
	lastEnteredDate = lastEnteredDate + timedelta(days=1)
	dateString = lastEnteredDate.strftime('%m/%d/%y') #get lastEnteredDate as a string
	outputSheet.cell(row = 1, column = col, value = dateString) 
	
	#copy the original scheduled column over.
	curRow = 2
	while curRow <= maxRow:
		originalCell = outputSheet.cell(row = curRow, column = col-7)
		newCell = outputSheet.cell(row = curRow, column = col)
		newCell.value = originalCell.value
		curRow += 1
	


print ("output saved to HumanityScheduleImport.xlsx")
outputWb.save('HumanityScheduleImport.xlsx')