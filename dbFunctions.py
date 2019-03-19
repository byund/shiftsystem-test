import MySQLdb
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import daySchedulingFunctions
import AuxiliaryFunctions
import re
import configparser
import os, shutil

def initializeEmployeeInfoTable(database, workerList): #takes in all the workers and makes a table out of them
	#initialize cursor
	cursor = database.cursor()
	
	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()
	
	#check if employee table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = %s "
		"AND table_name = 'employeeInfo'"
		"LIMIT 1", dbName)


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table of employees!")

	else: #if your query did not return a row, then the table does not exist.
		#first, create the table.
	#
		cmd = ("CREATE TABLE employeeInfo "
				"(id SMALLINT NOT NULL AUTO_INCREMENT, "
				"firstName VARCHAR(50), "	#Table parameter one: Names must be within 50 characters in length
				"lastName VARCHAR(50), "
				"username VARCHAR(25), "	#Check with Dave on maximum size a username can be to be efficient...also doublecheck if varchar is the best type for a name.		
				"colleagueID MEDIUMINT, "
				"phoneNumber VARCHAR(15), "
				"classYear YEAR(4), "
				"PRIMARY KEY (id))")

		cursor.execute(cmd)
		#We set employee IDs to start from 1001. 2000s are reserved for Admins.
		cursor.execute("ALTER TABLE employeeInfo AUTO_INCREMENT = 1001")
		
		#now we populate the employee table
		print ("Populating employees...")
		insertCmd = ("INSERT INTO employeeInfo (firstName, lastName, username, colleagueID)"
					"VALUES (%s, %s, %s, %s)"
			)

		for key in workerList: #remember, workerList is a dictionary of workers
			
			workerTuple = AuxiliaryFunctions.getStudent(workerList, key)
			print (workerTuple)	
			cursor.execute(insertCmd, workerTuple)

		#always close the cursor. Always commit all changes to the database.
		cursor.close()
		database.commit()



def initializeAdmins(database):
	cursor = database.cursor()
	
	adminFile = os.path.join(os.path.dirname(__file__), 'admins.txt') #Look up why this path patching is necessary.
	insertCmd = "INSERT INTO employeeInfo (id, firstName, lastName, username, colleagueID) VALUE (%s, %s, %s, %s, %s)"
	with open(adminFile) as f:
		content = f.readlines()
		for line in content:
			adminInfo = tuple(line.split(",")) #Get all info.
			print (adminInfo)
			cursor.execute(insertCmd, adminInfo)
	cursor.close()
	database.commit()



def initializeShiftList(database, startDate, endDate): 
	# #pass in startDate and endDate as Python dateTime objects
	#initialize cursor
	cursor = database.cursor()
	#print ("hello")

	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()
	
	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = %s "
		"AND table_name = 'ShiftList'"
		"LIMIT 1", (dbName))

	#print(cursor._last_executed)
	#print (cursor.rowcount)
	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table of shifts!")
	
	else: #Now we create the table if none exists
		cmd = ("CREATE TABLE ShiftList "
			"(id SMALLINT NOT NULL AUTO_INCREMENT PRIMARY KEY, "
			"startTime TIME,"
			"endTime TIME, "
			"Location VARCHAR(25) NOT NULL DEFAULT 'CMC', "
			"Date DATE NOT NULL, " #A DATE must be input in the form YYYY/MM/DD
			"endDate DATE, "
			"Day VARCHAR(15) )")
		

		cursor.execute(cmd)
		cursor.close()




		

		currentDate = startDate
	
	
		#having created the table, we populate the table. (NOTE: For ease of readability, I've shifted the methods that create the shifts for each individual day off to another file)
		print ("Now building Schedule of Shifts....")
		while (currentDate <= endDate):
			if currentDate.strftime("%A") == "Sunday" or currentDate.strftime("%A") == "Saturday":
				daySchedulingFunctions.buildWkndSchedule(database, currentDate)
				#print ("GROUNDBEEF")
			elif currentDate.strftime("%A") == "Monday" or currentDate.strftime("%A") == "Wednesday":
				#print ("HELLO")
				daySchedulingFunctions.buildMonWedSchedule(database, currentDate)
			elif currentDate.strftime("%A") == "Tuesday" or currentDate.strftime("%A") == "Thursday":
				#print ("BEEP")
				daySchedulingFunctions.buildTueThuSchedule(database, currentDate)

			elif currentDate.strftime("%A") == "Friday":
				#print ("MEEP")
				daySchedulingFunctions.buildFriSchedule(database, currentDate)

			currentDate = currentDate + timedelta(days = 1)
			
		#We remove all shifts prior to 8AM on the first day and all shifts after 5pm on the last day.
		cursor = database.cursor()
		headCmd = "DELETE FROM ShiftList WHERE date = %s AND startTime < '08:00'"
		tailCmd = "DELETE FROM ShiftList WHERE date = %s AND startTime > '21:00'"
		cursor.execute(headCmd, (startDate,))
		cursor.execute(tailCmd, (endDate,))


		#commit all changes to database
		cursor.close()
		database.commit()


		

		



def initializeShiftEmployeeLinker(database, listofWorkers, scheduleSheet, startDate, endDate):
	#we're assuming that shiftdb and employeedb have already been created.

	#initialize cursor
	cursor = database.cursor()
	
	#currentDate will be the closest Monday before the start date- this method needs to start on a monday to properly cycle through. Later, we will simply dispose of all shifts that are prior to the actual start date.
	currentDate = startDate - timedelta(days = startDate.weekday()) 
	
	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()

	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = %s "
		"AND table_name = 'shiftEmployeeLinker'"
		"LIMIT 1", dbName)


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table linking employees and shifts")

	else:
		#shift linker table does not exist, we create it first.
	
		cursor.execute("CREATE TABLE shiftEmployeeLinker "
			"(employeeID SMALLINT NOT NULL, "
			"shiftID SMALLINT NOT NULL, "
			"checkedIn BOOL NOT NULL DEFAULT 0, "
			"checkinTime TIME,"
			"subRequested BOOL NOT NULL DEFAULT 0, "
			"subFilled BOOL NOT NULL DEFAULT 0, "
			"FOREIGN KEY (employeeID) REFERENCES employeeInfo(id) ON DELETE CASCADE, "
			"FOREIGN KEY (shiftID) REFERENCES ShiftList(id) ON DELETE CASCADE, "
			"PRIMARY KEY (employeeID, shiftID))"

			)


		workerName = ""
		#now to populate the shiftlinker table. We will cycle through the entire Schedule Spreadsheet and match each worker to their shift.
		print ("Assigning workers to shifts...")
		
		timeRegex = re.compile('\d?\d:\d\d') #This regex matches any string that contains a Time (xx:xx)
		shiftRegex = re.compile('\dh')
		miscRegex = re.compile('[^a-zA-z.\s]') #This regex matches any string that contains a number i.e a non-name cell.

		while currentDate <= endDate:
			shiftTime = "8:00" #The helpdesk starts bright and early!
			Day = "Monday"
			
			AMBoolean = True

			for col in range(1,15): #this corresponds to columns A through N
				#print (currentDate)

				if (col % 2 == 0): #Any even numbered column is a library shift column.
					currentDate = currentDate + timedelta(days = -1) #flip back the date
					#print ("flipping")
				for row in range(2,53): #this corresponds to the maximum row that the schedule spreadsheet currently has. Modify as necessary. #THIS DOESN'T HIT THE LAST SHIFT ON SUNDAY NIGHT IN THE LIBE.
									
					cellValue = scheduleSheet.cell(row = row, column = col).value

					#print (cellValue)
					if cellValue != None: 
						cellValue = str(cellValue) #make sure you're dealing with a string.
						
						if timeRegex.search(cellValue) == None and shiftRegex.search(cellValue) == None: #You're not dealing with a time, or any cell that has a number in it.

							if miscRegex.search(cellValue) == None:
								#Day = AuxiliaryFunctions.getDay(row, col) #Get the day of the shift as a string.
								Day = currentDate.strftime("%A")
								date = currentDate.strftime("%Y/%m/%d")
								
								#get date of the shift as well

								location = AuxiliaryFunctions.getLocation(col) #get the location of the shift.
								#shiftTime = AuxiliaryFunctions.getShiftTime(row, col)


								#now to get the name of the worker (and thus their identity from the employeeTable)
								
								workerName = AuxiliaryFunctions.getStudent(listofWorkers, cellValue)

								#print (workerName)
								#print ((date, shiftTime))
								# print (date)
								# print ()

								employeeQuery = ("SELECT id " 	#TODO: FIGURE OUT AN ELEGANT WAY TO SOPHIA PROCESS
										"FROM employeeInfo "
										"WHERE firstName = %s "
										"AND lastName = %s"
										"AND username = %s"
										)

								
								
								#workerName is a tuple of the form (firstName, lastName, username, colleagueID)
								
								cursor.execute(employeeQuery, workerName[:3])
								result = cursor.fetchone()
								#for some reason cursor returns results as a tuple so....
								if not result:
									print ("Didn't find a result for this tuple")
									print (workerName)
									continue

								employeeid = result[0]

								#print ("Employee: " + str(employeeid))
								
								
								#If we have a worker, we must also have a shift!
								shiftQuery = ("SELECT id " 
										"FROM ShiftList "
										"WHERE Day = %s "
										"AND startTime = %s "
										"AND Location = %s "
										"AND Date = %s"
										)
								cursor.execute(shiftQuery, (Day, shiftTime, location, date))
								result = cursor.fetchone()
								if result != None:
									shiftid = result[0]
								else:
									print ("Couldn't find specified shift")
									print ((Day, shiftTime, location, date))
									continue
								
								#having gotten an employeeID and a shiftID, we append things to the shiftlinker table.
								inputQuery = ("INSERT INTO shiftEmployeeLinker (employeeID, shiftID)"
												"VALUES (%s, %s)")

								values = (employeeid, shiftid)

								cursor.execute(inputQuery, values)
								print(cursor._last_executed)

						else: #If you've hit a cell that has a number in it, you need to extract the shift time.
							#first, store the previous shift time. 
							prevShiftTime = datetime.strptime(shiftTime, "%H:%M")
							if timeRegex.search(cellValue) == None: #if you don't have a time match, you have hit a cell that tells you how long a shift is. 
							#you need to switch to the lefthand column
								cellValue = str(scheduleSheet.cell(row = row, column = col-1).value)

							shiftTime = timeRegex.search(cellValue).group(0)

							if "12" in shiftTime or shiftTime == "1:05": #the or statement is there just to catch the friday 4a shift, which has a weird time and won't toggle the boolean.
								#flip AM/PM if there's a 12 in the hour.
								AMBoolean = not AMBoolean


							#Next, convert shift time to 24hour format.
							if AMBoolean:
								shiftTime = shiftTime + "AM"
							else:
								shiftTime = shiftTime + "PM"

							
							shiftTime = datetime.strptime(shiftTime, "%I:%M%p")
							


							#deal with yet another edge case on weekends... 
							if (Day == "Friday" or Day == "Saturday") and shiftTime == datetime.strptime("20:00", "%H:%M"):
								AMBoolean = True

								
							
							#We check if the new shiftTime is earlier than the old shiftTime. If it's earlier, this indicates that we've flipped a Day.
							if shiftTime < prevShiftTime:
								currentDate = currentDate + timedelta(days = 1)

							#Finally, output ShiftTime as a string from its datetime representation.
							shiftTime = shiftTime.strftime("%H:%M")


	#cursor.close()
	

	#Finally, once the method has finished populating the table, we go into the shiftdb and drop all shifts that are prior to the startdate.

	cursor.execute ("DELETE FROM ShiftList WHERE date < %s", (startDate.strftime("%Y-%m-%d"),)) 
	
	cursor.close()
	database.commit()


def initializeSubbedShiftLinker(database):
	#This will create the table that manages subbed Shifts.
	cursor = database.cursor()

	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()

	#check if table exists.
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = %s "
		"AND table_name = 'subbedShifts'"
		"LIMIT 1", dbName)

	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a table linking Subbing employees and shifts")

	else:
		#initialize the table!
		cursor.execute("CREATE TABLE subbedShifts "
			"(shiftID SMALLINT NOT NULL, "
			"origEmployeeID SMALLINT NOT NULL, "
			"subEmployeeID SMALLINT NOT NULL, "
			"checkinTime TIME,"
			"FOREIGN KEY (shiftID) REFERENCES ShiftList(id) ON DELETE CASCADE, "
			"FOREIGN KEY (origEmployeeID) references employeeInfo(id) ON DELETE CASCADE, "
			"FOREIGN KEY (subEmployeeID) references employeeInfo(id) ON DELETE CASCADE, "
			"FOREIGN KEY (shiftID, origEmployeeID) REFERENCES shiftEmployeeLinker(shiftID, employeeID) ON DELETE CASCADE, "
			"PRIMARY KEY (shiftID, subEmployeeID))"

			)




def initializeDB(startDate): #Creates the database for the term schedule but does not populate it.
	#here, database is the overall database of schedules, and termdbname will be the termly schedule.
	config = configparser.ConfigParser()
	config.read('dblogin.ini')
	#kwargs.get[adminTrue, True] #returns true value
	host = config['LOGINADMIN']['host']
	user = config['LOGINADMIN']['user']
	password = config['LOGINADMIN']['password']
	
	database = MySQLdb.connect(host = host,
								user = user,
								passwd = password)
	print ("Initializing Database....")
	cursor = database.cursor()
	#name = (termDBname,) #a one-tuple of names!
	#name = input("What do you want to name this Table Schedule? ")
	if startDate.month >= 3 and startDate.month < 6:
		termDBname = "Spring" + str(startDate.year)
	elif startDate.month >= 9 and startDate.month < 12:
		termDBname = "Fall" + str(startDate.year)
	elif startDate.month >= 1 and startDate.month < 3:
		termDBname = "Winter" + str(startDate.year)
	else:
	#	print ("You don't even know who's working for you next year! Nice Try.")
		termDBname = "Summer" + str(startDate.year)
		
	#check if the database already exists.
	cmd = ("SHOW DATABASES LIKE '" + termDBname + "'")
	cursor.execute(cmd)
	#print (termDBname) 

	if (cursor.fetchone() == None):
		#if there is no database.
		cmd = ("create database if not exists " + termDBname) #YOU CANNOT PARAMETRIZE TABLE NAMES...BEWARE OF SQL INJECTION ATTACKS
	#print (cmd)
		try:
			cursor.execute(cmd)
			print("Database '" + termDBname + "' successfully created.")
			print("Continuing the Song that ends the world")
		except :

			print (cursor._last_executed)

		
		#Create a configparser to manipulate INI files.
		config = configparser.ConfigParser()
		
		#check if a previous currentDB.ini exists. If it does, move it to the databases folder and rename it.
		if os.path.isfile('currentDB.ini'):
			print ("Found preexisting currentDB initialization file. Moving to databases/...") 
			config.read('currentDB.ini')
			dbName = config['DBINFO']['database']
			dst_dir = "databases/" + dbName + ".ini"
			shutil.move('currentDB.ini', dst_dir)
			print ("Successfully moved .ini file. Cataclysm draws ever closer...")	
		else:
			print ("Ho")
		#create an INI file that stores the login credentials for this new database. See https://docs.python.org/3/library/configparser.html
		config['DBINFO'] = {
		'database':termDBname, 'adminDatabase':termDBname

	#todo- put in a clause to add startdate as well.
		}
	

		with open ('currentDB.ini', 'w') as configFile:
			config.write(configFile)

		#TODO: Make it so that everytime you create a new database, it saves the old login credentials and moves them to another folder.

	else:
		print ("the database for this term already exists. Continuing this process will accomplish nothing. I suggest you flee while you still have the chance.")
		database.close()
		return False

	

def initializeViews(db):

	
	cursor = db.cursor()
	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()

	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.tables "
		"WHERE table_schema = %s "
		"AND table_name = 'quickReportView'"
		"LIMIT 1", dbName)


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a quickReportView.")

	else:

		cmd = ("CREATE VIEW quickReportView AS "
		"SELECT employeeInfo.id as employeeID, ShiftList.id as shiftID, employeeInfo.firstName, employeeInfo.lastName, ShiftList.date, ShiftList.startTime, shiftEmployeeLinker.checkinTime "
		"FROM employeeInfo "
		"INNER JOIN shiftEmployeeLinker "
		"ON employeeInfo.id = shiftEmployeeLinker.employeeID "
		"INNER JOIN ShiftList "
		"ON shiftEmployeeLinker.shiftID = ShiftList.id "
		)
	
	
	#"CREATE VIEW allShifts AS"
	#"SELECT ShiftList.checkinTime AS startTime, ShiftList2.checkinTime AS endTime "
	#"FROM ShiftList "
	#"INNER JOIN ShiftList AS ShiftList2 "
	#"ON ShiftList2.id = ( "
	#"SELECT ShiftList3.id "
	#"FROM ShiftList ShiftList3 "
	#"WHERE ShiftList.Location = ShiftList3.Location "
	#"AND ShiftList3.id > ShiftList.id "
	#"LIMIT 1)"
		# "UNION"
		# "SELECT subbedShifts.subEmployeeID AS workingEmployee, subbedShifts.shiftID, ShiftList.checkinTime AS startTime, shiftList2.checkinTime as endTime "
		# "FROM ShiftList "
		# "INNER JOIN ShiftList as ShiftList2 "
		# "ON ShiftList2.id = ("
		# 	"SELECT ShiftList3.id "
		# 	"FROM ShiftList ShiftList3"
		# 	"WHERE ShiftList.Location = ShiftList3.Location "
		# 	"AND ShiftList3.id > ShiftList.id "
		# 	"LIMIT 1)"

		cursor.execute(cmd)
	cursor.close()
	db.commit()


def dbLogIn(**kwargs): #reads host, user, password from dblogin.ini and logs into mysql server. Selects a particular server based on current configuration seen in currentDB.ini
	#TODO: add flag to log in as admin or not.
	config = configparser.ConfigParser()
	loginFile = os.path.join(os.path.dirname(__file__), 'dblogin.ini')
	config.read(loginFile)
	#config.read('dblogin.ini')
	isAdmin = kwargs.get("admin", False) #returns False value unless isAdmin is explicitly set.
	host = config['LOGINADMIN']['host']
	user = config['LOGINADMIN']['user']
	password = config['LOGINADMIN']['password']

	currentDBFile = os.path.join(os.path.dirname(__file__), 'currentDB.ini') #Look up why this path patching is necessary.
	config.read(currentDBFile)	

	#if the database is manually specified when calling the function...
	if kwargs.get("database"):
		database = kwargs["database"]
	else:
		if isAdmin:
			database = config['DBINFO']['adminDatabase']
		else:	
			database = config['DBINFO']['database']

	db = MySQLdb.connect(host = host,
								user = user,
								passwd = password,
								db = database)

	#print ("Now logged into " + database)


	return db


def initializeStoredProcedures(db):
	cursor = db.cursor()
	#get currentDatabase Name
	cursor.execute("SELECT DATABASE()")
	dbName = cursor.fetchone()

	#check if shift table exists
	cursor.execute("SELECT * " 
		"FROM information_schema.routines "
		"WHERE routine_schema = %s "
		"AND routine_name = 'quickReport'"
		"LIMIT 1", dbName)


	if (cursor.rowcount != 0): #If your query returned a row, then the table exists
		print ("There is already a quickReport Stored Procedure.")

	else:
	
		cursor.execute("""
	CREATE PROCEDURE quickReport (IN currTime datetime, IN startTime datetime)
	BEGIN
		SELECT * FROM quickReportView WHERE CAST(CONCAT(date, ' ', quickReportView.startTime) as datetime) < currTime
		AND CAST(CONCAT(date, ' ', quickReportView.startTime) AS datetime) >= startTime;
	END;""")

	
		cursor.execute("""
	CREATE PROCEDURE quickWorkerSummary (IN currTime datetime, IN startTime datetime, IN employeeID smallint)
	BEGIN
		SELECT * FROM quickReportView WHERE CAST(CONCAT(date, ' ', quickReportView.startTime) as datetime) < currTime 
		AND CAST(CONCAT(date, ' ', quickReportView.startTime) AS datetime) >= startTime
		AND employeeID = quickReportView.employeeID;
	END;""")

	cursor.close()

	return	

def initializeFallFirstWeek(db):
	#if it's fall term, tweaks ShiftList so that the scheduled shifts for Monday and Friday conform to the weird schedule at the beginning of every year.
	print ("Tweaking for Fall 1st Week...")
	#initialize Cursor..
	cursor = db.cursor()
	
	mondayCMD = """UPDATE ShiftList SET startTime = %s, endTime = %s WHERE day = "Monday" and startTime = %s LIMIT 2"""
	mondayParams = [
("08:00", "09:25", "08:00"),
("09:25", "10:25", "09:45"),
("10:25", "11:25", "11:05"),
("11:25", "12:25", "12:25"),
("12:25", "13:25", "13:45"),
("13:25", "14:25", "15:05"),
]

	cursor.executemany(mondayCMD, mondayParams)
	print ("Tweaked Monday!")
	
	fridayCMD = ("UPDATE ShiftList SET startTime = %s, endTime = %s WHERE day = 'Friday' and startTime = %s LIMIT 2")
	#First delete the convo shifts... 
	cursor.execute("DELETE FROM ShiftList WHERE startTime = '10:45' AND endTime = '11:55' AND day = 'Friday' LIMIT 2")
	fridayParams = [
("08:00", "09:45", "08:00"),
("09:45", "11:05", "09:35"),
("11:05", "12:25", "11:55"),
("12:25", "13:45", "13:05"),
("13:45", "15:05", "14:15"),
("15:05", "16:30", "15:25"),
]


	cursor.executemany(fridayCMD, fridayParams)
	
	print ("Tweaked Friday!")
	db.commit()
	print ("Schedule successfully Tweaked! Seen enough 'Tweaks' yet?")

def accessDB(function, *args, isAdmin=None, **kwargs):
	
	if not isAdmin:
		isAdmin = False;	
	
	db = dbLogIn(admin = isAdmin)
	#print ("successfully logged in")
	
	result = function(db, *args, **kwargs)

	
	db.close()
	#print ("closed DB")
	
	return result

def main():

	filename = input("Place ITS Schedule Spreadsheet in the same folder as this program, then type its filename here (e.g Schedule.xlsx): ")
	#wb = load_workbook("Spring Schedule.xlsx")
	wb = load_workbook(filename)
	scheduleSheet = wb['CarlTech Schedule']
	staffSheet = wb['CarlTech Staff List']
	workerList = AuxiliaryFunctions.getListOfWorkers(wb)

	dateholder = input("When does the term start? (mm/dd/yyyy): ")
	startDate = datetime.strptime(dateholder, "%m/%d/%Y")
	dateholder = input("When does the term end? (mm/dd/yyyy): ")
	endDate = datetime.strptime(dateholder, "%m/%d/%Y")
	
	print ("Now beginning the Song that will herald the end of days.....")
	initializeDB(startDate)
	db = dbLogIn() #remember to close the database after you're done!


	initializeShiftList(db, startDate, endDate)
	initializeEmployeeInfoTable(db, workerList)
	initializeAdmins(db)
	initializeShiftEmployeeLinker(db, workerList, scheduleSheet, startDate, endDate)
	initializeSubbedShiftLinker(db)
	initializeViews(db)
	initializeStoredProcedures(db)
	
	if startDate.month >= 9 and startDate.month < 12: #if it's fall term....
		initializeFallFirstWeek(db)
	db.close()

	
	print ("The Song is now over. End of World is now approximately 23.456821359E20 millennia closer. You may now unfasten your seatbelt. Thank you for using ShiftSystem TM, a product of Appachar Industries.")
	print("It is now safe to turn off your computer.")

if __name__ == "__main__":
	main()
